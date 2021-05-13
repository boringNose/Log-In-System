import re
import os
from openpyxl import Workbook, load_workbook
# from getpass import getpass


def validate_username(name):
    """
        validate_username(name)
        Purpose: validates name for the user
        User name cannot be:
            - less than 6 and more than 20 characters
            - cannot contain any non-word characters
            - must not start with digit or underscore
    """
    while len(name) < 6 or len(name) > 20:
        print("Username must be in between 6 and 20 characters.")
        name = input("Please enter a valid username: ").strip()
    while True:
        match1 = re.findall(r"\W", name)     # Returns a match for non-word characters
        match2 = re.findall(r"^\d|^_", name)  # Returns a match if username starts with a digit
        if len(match1) == 0 and len(match2) == 0:
            return name
        else:
            if len(match1) != 0:
                print("Username can only contain Alphabets(a-zA-Z), Digits(0-9) or Underscore(_).")
            else:
                print("Username cannot start with a Digit or underscore(_).")
            name = input("Please enter a valid username: ").strip()


def validate_password(pwd):
    flag = True
    while True:
        validate_dict = {"uppercase": 0, "lowercase": 0, "digit": 0, "special character": 0}
        for x in pwd:
            if x.isupper():
                validate_dict["uppercase"] += 1
            elif x.islower():
                validate_dict["lowercase"] += 1
            elif x.isdigit():
                validate_dict["digit"] += 1
            else:
                validate_dict["special character"] += 1

        for x, y in validate_dict.items():
            if y == 0:
                print(f"Password must contain at least one {x} in it.")
                flag = False
                break
        if flag and len(pwd) >= 8:
            return pwd
        else:
            if len(pwd) < 8:
                print("Password should be minimum 8 characters long.")
            pwd = input("Enter a valid password: ").strip()
            # pwd = getpass("Enter a valid password: ").strip()
            flag = True
            validate_dict.clear()


def get_cells_and_values(counter):
    cell_a = "A" + str(counter)
    cell_b = "B" + str(counter)
    return [(cell_a, ws[cell_a].value), (cell_b, ws[cell_b].value)]


def add_user(name):
    counter = 2     # initialized with 2 because first row contains headings (username, password) for the file
    while True:
        cells_values = get_cells_and_values(counter)  # get cells and values for column A and B as a list of tuples
        if cells_values[0][1] == name:  # checks if name entered by user matches with any name in the file
            print(f"User cannot be added. {name} already exists.")
            break
        elif cells_values[0][1] is None:    # checks if cell value is empty
            pwd = validate_password(input("Enter password: ").strip())
            ws[cells_values[0][0]] = name
            ws[cells_values[1][0]] = pwd
            print(f"{name} successfully added in the system.")
            break
        counter += 1
    wb.save(filename)


def login(name):
    counter = 2     # initialized with 2 because first row contains headings (username, password) for the file
    cells_values = get_cells_and_values(counter)    # get cells and values for column A and B as a list of tuples
    user_names = []
    user_passwords = []
    while cells_values[0][1] is not None:
        user_names.append(cells_values[0][1])   # store all user names in a list
        user_passwords.append(cells_values[1][1])   # store all user passwords in a list
        counter += 1
        cells_values = get_cells_and_values(counter)
    else:
        if name not in user_names:
            print(f"{name} is not registered in the system.")
        else:
            index = user_names.index(name)
            pwd = input("Log in password: ").strip()
            while len(pwd) < 8:
                pwd = input("Password cannot be less than 8 characters.\nEnter password to login: ")
            else:
                if pwd == user_passwords[index]:
                    print(f"{name} successfully logged into the system.")
                else:
                    print(f"Wrong Password.")


def user_selection():
    selection = 0
    while True:
        try:
            choice = int(input("Select your choice: \n1. Register\n2. Log In\nEnter choice here: "))
        except ValueError as e:
            print("\n", e)
            continue
        else:
            if choice == 1 or choice == 2:
                selection = choice
                break
            else:
                print("\nChoose either 1 or 2!")
    return selection


filePath = r"E:\login_system\login_spreadsheet.xlsx"
filename = "login_spreadsheet.xlsx"

# checks if excel file already exists or not
if os.path.exists(filePath):
    wb = load_workbook(filename)
    ws = wb.active
    for sheet in wb.sheetnames:
        print(sheet.title())
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Users' list"
    ws["A1"] = "Username"
    ws["B1"] = "Password"

exit_system = ""
user_choice = user_selection()  # asks user to register or log in

# loop forever until the user presses 'n' or forcefully stops the execution of program
while exit_system != "n":
    if user_choice == 1:
        print("Welcome! Please enter your details to register.")
        username = validate_username(input("Enter username: ").strip())   # Removes any leading or trailing spaces
        add_user(username.lower())    # store users with valid credentials in an excel file
        # password = validate_password(getpass("Enter password: ").strip())
        # To use getpass go to edit configurations and select emulate terminal in output console
    else:
        print("Enter your details to login.")
        username = input("Log in username: ").strip()
        login(username.lower())     # allow users with valid credentials to log into the system
    exit_system = input("Do you want to continue? (Press n to quit): ")
    if exit_system != "n":
        user_choice = user_selection()

wb.save(filename)   # saves the workbook
wb.close()  # closes the workbook
# print(validate_username.__doc__)      # prints docstring





